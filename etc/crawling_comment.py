import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time

wb = load_workbook('./연도별 박스오피스 순위.xlsx')
ws = wb.active

driver = webdriver.Chrome('./chromedriver')
url = 'https://movie.naver.com/'
driver.get(url)
time.sleep(1)


for r in range(4, ws.max_row+1):
# for r in range(491, 494):
    if r >= 491:
        x_path = '//*[@id="ipt_tx_srch"]'
        searchbox = driver.find_element_by_xpath(x_path)
        searchbox.click()

        title = ws.cell(row=r, column=2).value
        element = driver.find_element_by_id("ipt_tx_srch")
        element.send_keys(title)
        element.send_keys(Keys.ENTER)

        time.sleep(0.5)
        element1 = driver.find_element_by_xpath('//*[@id="old_content"]/ul[2]/li[1]').find_element_by_class_name("result_thumb")
        element1.click()

        time.sleep(0.5)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        result = soup.select_one('.score_result')
        
        try:
            liTag = result.select_one('#content > div.article > div.section_group.section_group_frst > div:nth-child(5) > div:nth-child(2) > div.score_result > ul > li:nth-child(2)')

            # print(liTag)
            score = liTag.select_one('em').get_text()
            comment = liTag.select_one('p').get_text()
            comment = comment.strip()
            nickname = liTag.select_one('dl > dt > em > a > span').get_text()

            ws.cell(row=r, column=4).value = score
            ws.cell(row=r, column=5).value = nickname
            ws.cell(row=r, column=6).value = comment

            print(nickname, comment)
        
        except:
            pass

    else:
        x_path = '//*[@id="ipt_tx_srch"]'
        searchbox = driver.find_element_by_xpath(x_path)
        searchbox.click()

        title = ws.cell(row=r, column=2).value
        element = driver.find_element_by_id("ipt_tx_srch")
        element.send_keys(title)
        element.send_keys(Keys.ENTER)

        time.sleep(0.5)
        element1 = driver.find_element_by_xpath('//*[@id="old_content"]/ul[2]/li[1]').find_element_by_class_name("result_thumb")
        element1.click()

        time.sleep(0.5)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        result = soup.select_one('.score_result')
        
        try:
            liTag = result.select_one('#content > div.article > div.section_group.section_group_frst > div:nth-child(5) > div:nth-child(2) > div.score_result > ul > li:nth-child(1)')

            # print(liTag)
            score = liTag.select_one('em').get_text()
            comment = liTag.select_one('p').get_text()
            comment = comment.strip()
            nickname = liTag.select_one('dl > dt > em > a > span').get_text()

            ws.cell(row=r, column=4).value = score
            ws.cell(row=r, column=5).value = nickname
            ws.cell(row=r, column=6).value = comment

            print(nickname, comment)
        
        except:
            pass

wb.save('test.xlsx')
driver.quit()

    
