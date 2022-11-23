# kobis에서 찾은 데이터에 포스터 & 요약내용 추가하기

import requests
import json
from openpyxl import load_workbook

wb = load_workbook('./연도별 박스오피스 순위.xlsx')
ws = wb.active

API_KEY='bd7f98121a9d0436318b3160e3374695'

def get_movie_datas():
    total_data = []

    for r in range(5, ws.max_row+1):
    # for r in range(3, 4):
        movie_title_en = ws.cell(row=r, column=8).value

        url = f'https://api.themoviedb.org/3/search/movie?api_key={API_KEY}&query={movie_title_en}&language=ko-KR'

        movies = requests.get(url).json()
        # festival_name = ws.cell(row=r, column=3).value

        for movie in movies['results']:   
            release_date = movie['release_date'][:4]
            print(movie_title_en)
            fields = {
                    'title':  ws.cell(row=r, column=1).value,
                    'director': ws.cell(row=r, column=2).value,
                    'year': ws.cell(row=r, column=3).value,
                    'nation': ws.cell(row=r, column=4).value,
                    'titleEn': ws.cell(row=r, column=8).value,
                    'genre': ws.cell(row=r, column=6).value,
                    'adult': ws.cell(row=r, column=7).value,
                    'vote_average': movie['vote_average'],
                    'overview': movie['overview'],
                    'poster_path': movie['poster_path'],
                    }
        
            data = {
                    "model": "movies.award_Movie",
                    "pk": movie['id'],  
                    "fields": fields
                    }
        
            total_data.append(data)
            break

    with open("award_data.json", 'w', encoding='utf-8') as w:
        json.dump(total_data, w, indent=" ", ensure_ascii=False)

get_movie_datas()
