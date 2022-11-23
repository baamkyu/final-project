import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By

wb = load_workbook('./연도별 박스오피스 순위.xlsx')
ws = wb.active

driver = webdriver.Chrome('/path/to/chromedriver_win32')
url = 'https://movie.naver.com/'
driver.get(url)
time.sleep(0.2)



# for r in range(5, 10):
for r in range(5, 1004):
  x_path = '//*[@id="ipt_tx_srch"]'
  # searchbox = driver.find_element_by_xpath(x_path)
  searchbox = driver.find_element(By.XPATH, x_path)
  searchbox.click()


  title = ws.cell(row=r, column=1).value
  # element = driver.find_element_by_id("ipt_tx_srch")
  element = driver.find_element(By.ID, "ipt_tx_srch")
  element.send_keys(title)
  element.send_keys(Keys.ENTER)
  time.sleep(0.2)
  try:
  # element1 = driver.find_element_by_xpath('//*[@id="old_content"]/ul[2]/li[1]').find_element_by_class_name("result_thumb")
    element1 = driver.find_element(By.XPATH, '//*[@id="old_content"]/ul[2]/li[1]').find_element(By.CLASS_NAME, "result_thumb")
    element1.click()

    time.sleep(0.2)
    html = driver.page_source
    soup = BeautifulSoup(html, 'html.parser')
    enTitle = soup.select_one('.h_movie2').get_text()
    print(enTitle)

    poster = soup.select_one('#content > div.article > div.mv_info_area > div.poster > a > img').get('src')
    # print(poster)

    director = soup.select_one('#content > div.article > div.mv_info_area > div.mv_info > dl > dd:nth-child(4) > p').get_text()


    ws.cell(row=r, column=8).value = enTitle
    ws.cell(row=r, column=10).value = director
    ws.cell(row=r, column=11).value = poster
  except:
    pass

wb.save('test.xlsx')
driver.quit()
